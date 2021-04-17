from openpyxl import load_workbook, Workbook
import os




def spreadsheet_automation():
    file_arr = []
    removed_dups = []
    finalworkbook = Workbook()
    finalsheet = finalworkbook.active
    list_dic = []
    email_list = []

    finalsheet.title = "RRR Mailing List"

    finalsheet['A1'] = "Name:"
    finalsheet['B1'] = "Email:"
    finalsheet['C1'] = "Phone:"
    finalsheet['D1'] = "Address 1:"
    finalsheet['E1'] = "Address 2:"

    local_folder = "/Programming/rrr/"
    for path, dirnames, filenames in os.walk(local_folder):
            file_arr.extend(os.path.join(name) for name in filenames)

    n = 2
    del file_arr[-n:]
    print(file_arr)

    for file in file_arr[0:len(file_arr) - 1]:
        workbook_file = file
        wb = load_workbook(filename = local_folder + workbook_file)
        if(file.endswith('.xlsx')):
            if 'Input Sheet' in wb.sheetnames:
                sheet_ranges = wb['Input Sheet']
                name = sheet_ranges['A3'].value
                phone = sheet_ranges['A6'].value
                email = sheet_ranges['A7'].value
                address1 = sheet_ranges['A4'].value
                address2 = sheet_ranges['A5'].value
                if email is not None and email not in email_list:
                    finalsheet.cell(column=1, row=finalsheet.max_row+1, value=name)
                    finalsheet.cell(column=2, row=finalsheet.max_row, value=email)
                    finalsheet.cell(column=3, row=finalsheet.max_row, value=phone)
                    finalsheet.cell(column=4, row=finalsheet.max_row, value=address1)
                    finalsheet.cell(column=5, row=finalsheet.max_row, value=address2)
                    print(name, phone, email) 
                    email_list.append(email)
    finalworkbook.save('reliablelist.xlsx')



if __name__ == "__main__":
    spreadsheet_automation()
           